import io
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.user import User
from app.models.payment import Payment
from app.models.import_history import ImportHistory
from app.routes.auth import get_current_user

router = APIRouter(
    prefix="/api/payments",
    tags=["Payments"]
)

HEADER_MAP = {
    "Sub Order No": "sub_order_no",
    "Order Date": "order_date",
    "Dispatch Date": "dispatch_date",
    "Product Name": "product_name",
    "Supplier SKU": "sku",
    "Catalog ID": "catalog_id",
    "Order source": "order_source",
    "Live Order Status": "live_order_status",
    "Product GST %": "product_gst_percentage",
    "Listing Price (Incl. taxes)": "listing_price",
    "Quantity": "quantity",
    "Transaction ID": "transaction_id",
    "Payment Date": "payment_date",
    "Final Settlement Amount": "final_settlement_amount",
    "Price Type": "price_type",
    "Total Sale Amount (Incl. Shipping & GST)": "total_sale_amount",
    "Total Sale Return Amount (Incl. Shipping & GST)": "total_sale_return_amount",
    "Fixed Fee (Incl. GST)": "fixed_fee",
    "Warehousing fee (inc Gst)": "warehousing_fee",
    "Warehousing fee (Incl. GST)": "warehousing_fee",
    "Return premium (incl GST)": "return_premium",
    "Return premium (incl GST) of Return": "return_premium_return",
    "Meesho Commission Percentage": "commission_percentage",
    "Meesho Commission (Incl. GST)": "commission_amount",
    "Meesho gold platform fee (Incl. GST)": "gold_platform_fee",
    "Meesho mall platform fee (Incl. GST)": "mall_platform_fee",
    "Return Shipping Charge (Incl. GST)": "return_shipping_charge",
    "GST Compensation (PRP Shipping)": "gst_compensation",
    "Shipping Charge (Incl. GST)": "shipping_charge",
    "Other Support Service Charges (Excl. GST)": "other_support_service_charges",
    "Waivers (Excl. GST)": "waivers",
    "Net Other Support Service Charges (Excl. GST)": "net_other_support_service_charges",
    "GST on Net Other Support Service Charges": "gst_on_support_service_charges",
    "TCS": "tcs",
    "TDS Rate %": "tds_rate",
    "TDS": "tds",
    "Compensation": "compensation",
    "Claims": "claims",
    "Recovery": "recovery",
    "Compensation Reason": "compensation_reason",
    "Claims Reason": "claims_reason",
    "Recovery Reason": "recovery_reason"
}

MAX_FILE_SIZE = 50 * 1024 * 1024

def clean_numeric(val):
    if val is None: return None
    if isinstance(val, (int, float)): return Decimal(str(val))
    val_str = str(val).replace('₹', '').replace(',', '').replace('%', '').strip()
    if not val_str or val_str == '-' or val_str.lower() == 'null': return None
    try:
        return Decimal(val_str)
    except InvalidOperation:
        return None

def clean_int(val):
    if val is None: return None
    if isinstance(val, int): return val
    if isinstance(val, float): return int(val)
    val_str = str(val).replace(',', '').strip()
    if not val_str or val_str == '-': return None
    try:
        return int(float(val_str))
    except ValueError:
        return None

def clean_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    val_str = str(val).strip()
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            pass
    return None

@router.post("/upload")
async def upload_payments(
    file: UploadFile = File(...), 
    reporting_period: Optional[str] = Form(None),
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only XLSX files are allowed.")
    
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds the 50MB limit.")
    if len(contents) == 0:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail="Failed to parse XLSX file. It might be corrupt.")
        
    if "Order Payments" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook must contain a sheet named 'Order Payments'.")
        
    ws = wb["Order Payments"]
    
    header_row = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[2]]
    
    header_indices = {}
    for i, h in enumerate(header_row):
        if h in HEADER_MAP:
            if h not in header_indices:
                header_indices[h] = i

    missing_headers = []
    for req_header in HEADER_MAP:
        if req_header not in header_indices and req_header != "Warehousing fee (Incl. GST)":
            missing_headers.append(req_header)

    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Missing required headers: {', '.join(missing_headers)}")

    existing_payments_query = db.query(Payment.sub_order_no, Payment.transaction_id).filter(Payment.user_id == current_user.id).all()
    existing_payments = {f"{row[0]}_{row[1]}" for row in existing_payments_query if row[0] and row[1]}

    payments_to_insert = []
    total_rows = 0
    duplicate_rows = 0
    processed_in_batch = set()
    month_counts = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
            
        total_rows += 1
        
        row_data = {}
        for header, db_field in HEADER_MAP.items():
            if header in header_indices:
                idx = header_indices[header]
                if idx < len(row):
                    row_data[db_field] = row[idx]
                else:
                    row_data[db_field] = None

        sub_order_no = str(row_data.get("sub_order_no", "")).strip() if row_data.get("sub_order_no") else None
        transaction_id = str(row_data.get("transaction_id", "")).strip() if row_data.get("transaction_id") else None

        if not sub_order_no or not transaction_id:
            continue

        dup_key = f"{sub_order_no}_{transaction_id}"
        if dup_key in existing_payments or dup_key in processed_in_batch:
            duplicate_rows += 1
            continue
            
        processed_in_batch.add(dup_key)

        payment_date_val = clean_date(row_data.get("payment_date"))
        if payment_date_val:
            m_str = payment_date_val.strftime("%B %Y")
            month_counts[m_str] = month_counts.get(m_str, 0) + 1
        else:
            order_date_val = clean_date(row_data.get("order_date"))
            if order_date_val:
                m_str = order_date_val.strftime("%B %Y")
                month_counts[m_str] = month_counts.get(m_str, 0) + 1

        new_payment = Payment(
            user_id=current_user.id,
            sub_order_no=sub_order_no,
            transaction_id=transaction_id,
            live_order_status=str(row_data.get("live_order_status", "")).strip() if row_data.get("live_order_status") else None,
            product_name=str(row_data.get("product_name", "")).strip() if row_data.get("product_name") else None,
            sku=str(row_data.get("sku", "")).strip() if row_data.get("sku") else None,
            catalog_id=str(row_data.get("catalog_id", "")).strip() if row_data.get("catalog_id") else None,
            order_source=str(row_data.get("order_source", "")).strip() if row_data.get("order_source") else None,
            price_type=str(row_data.get("price_type", "")).strip() if row_data.get("price_type") else None,
            compensation_reason=str(row_data.get("compensation_reason", "")).strip() if row_data.get("compensation_reason") else None,
            claims_reason=str(row_data.get("claims_reason", "")).strip() if row_data.get("claims_reason") else None,
            recovery_reason=str(row_data.get("recovery_reason", "")).strip() if row_data.get("recovery_reason") else None,
            
            quantity=clean_int(row_data.get("quantity")),
            
            payment_date=payment_date_val,
            order_date=clean_date(row_data.get("order_date")),
            dispatch_date=clean_date(row_data.get("dispatch_date")),
            
            final_settlement_amount=clean_numeric(row_data.get("final_settlement_amount")),
            product_gst_percentage=clean_numeric(row_data.get("product_gst_percentage")),
            listing_price=clean_numeric(row_data.get("listing_price")),
            total_sale_amount=clean_numeric(row_data.get("total_sale_amount")),
            total_sale_return_amount=clean_numeric(row_data.get("total_sale_return_amount")),
            fixed_fee=clean_numeric(row_data.get("fixed_fee")),
            warehousing_fee=clean_numeric(row_data.get("warehousing_fee")),
            return_premium=clean_numeric(row_data.get("return_premium")),
            return_premium_return=clean_numeric(row_data.get("return_premium_return")),
            commission_percentage=clean_numeric(row_data.get("commission_percentage")),
            commission_amount=clean_numeric(row_data.get("commission_amount")),
            gold_platform_fee=clean_numeric(row_data.get("gold_platform_fee")),
            mall_platform_fee=clean_numeric(row_data.get("mall_platform_fee")),
            return_shipping_charge=clean_numeric(row_data.get("return_shipping_charge")),
            gst_compensation=clean_numeric(row_data.get("gst_compensation")),
            shipping_charge=clean_numeric(row_data.get("shipping_charge")),
            other_support_service_charges=clean_numeric(row_data.get("other_support_service_charges")),
            waivers=clean_numeric(row_data.get("waivers")),
            net_other_support_service_charges=clean_numeric(row_data.get("net_other_support_service_charges")),
            gst_on_support_service_charges=clean_numeric(row_data.get("gst_on_support_service_charges")),
            tcs=clean_numeric(row_data.get("tcs")),
            tds_rate=clean_numeric(row_data.get("tds_rate")),
            tds=clean_numeric(row_data.get("tds")),
            compensation=clean_numeric(row_data.get("compensation")),
            claims=clean_numeric(row_data.get("claims")),
            recovery=clean_numeric(row_data.get("recovery"))
        )
        payments_to_insert.append(new_payment)

    try:
        if payments_to_insert:
            db.add_all(payments_to_insert)
            db.commit()
    except Exception as e:
        db.rollback()
        
        try:
            detected_period = max(month_counts, key=month_counts.get) if month_counts else None
            final_period = reporting_period or detected_period
            
            failed_history = ImportHistory(
                user_id=current_user.id,
                import_type="payments",
                reporting_period=final_period,
                original_filename=file.filename,
                file_size=len(contents),
                total_rows=total_rows,
                inserted_rows=0,
                duplicate_rows=0,
                status="failed"
            )
            db.add(failed_history)
            db.commit()
        except Exception:
            db.rollback()
            
        raise HTTPException(status_code=500, detail=f"Database error during insert: {str(e)}")

    try:
        detected_period = max(month_counts, key=month_counts.get) if month_counts else None
        final_period = reporting_period or detected_period
        
        success_history = ImportHistory(
            user_id=current_user.id,
            import_type="payments",
            reporting_period=final_period,
            original_filename=file.filename,
            file_size=len(contents),
            total_rows=total_rows,
            inserted_rows=len(payments_to_insert),
            duplicate_rows=duplicate_rows,
            status="success"
        )
        db.add(success_history)
        db.commit()
    except Exception as history_error:
        db.rollback()

    return {
        "message": "Payments processed successfully",
        "total_rows": total_rows,
        "inserted_rows": len(payments_to_insert),
        "duplicate_rows": duplicate_rows,
        "reporting_period": final_period
    }
